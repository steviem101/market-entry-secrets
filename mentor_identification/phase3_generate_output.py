"""
Phase 3: Apply Recipe Y filter (archetype + score floor 14) and produce the
final mes_mentor_candidates.xlsx + .csv + supporting sheets.

Inputs:
  archetype_classification.csv  (from archetype_simulator.py)

Outputs:
  mes_mentor_candidates.xlsx
  mes_mentor_candidates.csv
"""
import re
from collections import Counter
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

ROOT = Path(__file__).parent
df = pd.read_csv(ROOT / "archetype_classification.csv", dtype=str, keep_default_na=False)

# Numeric columns
for col in ["total_score", "seniority_score", "relevance_score", "industry_score",
            "network_score", "persona_fit_score", "intl_founder_score",
            "linkedinFollowersCount"]:
    df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

# Boolean columns (CSV stores as 'True'/'False')
for col in ["is_founder", "already_in_google_sheet", "already_in_supabase",
            "is_mes_tagged", "mes_company_match", "any_archetype",
            "archetype_advisor", "archetype_csuite_sector",
            "archetype_trade_gov", "archetype_intl_founder",
            "has_scale_signal"]:
    df[col] = df[col].astype(str).str.lower().isin(["true", "1"])

# ---------------------------------------------------------------------------
# RE-CATEGORIZATION: score every category, pick the strongest signal.
# Replaces the original "first keyword wins" heuristic which over-rotated
# everyone with "growth strategy" or "business development" into sales.
# ---------------------------------------------------------------------------
CAT_DEFS = {
    "government_trade": {
        "title": [
            r"\btrade\s+commissioner\b", r"\bconsul[\s-]*general\b",
            r"\bconsul\b", r"\bcountry\s+(director|manager|head)\b",
            r"\binvestment\s+commissioner\b", r"\bambassador\b",
            r"\battache\b", r"\battaché\b", r"\bchief\s+representative\b",
            r"\b(commercial|economic|trade)\s+(officer|attache|attaché|counsellor|secretary)\b",
            r"\blanding\s+pad\b", r"\btrade\s+(adviser|advisor)\b",
            r"\bsenior\s+adviser\b.{0,40}(trade|investment|austrade)",
            r"\beconomic\s+development\b",
        ],
        "company": [
            r"\baustrade\b", r"\bnzte\b", r"\binvestment\s+nsw\b",
            r"\b(embassy|consulate)\b", r"\bchamber\s+of\s+commerce\b",
            r"\btrade\s+(council|commission|agency|office|and|&)\b",
            r"\benterprise\s+(ireland|singapore)\b",
            r"\bida\s+ireland\b", r"\bkotra\b", r"\bjetro\b",
            r"\bbusiness\s+(france|sweden)\b",
            r"\b(catalonia|basque)\s+trade\b",
            r"\bglobal\s+affairs\s+canada\b",
            r"\binvest(ment)?\s+\w+\b", r"\bdepartment\s+for\s+business\b",
            r"\bexport\s+(council|finance|development)\b",
            r"\btrade\s+development\b", r"\bcouncil\b",
        ],
        "industry": ["international trade", "government administration",
                     "government relations", "international affairs", "diplomatic"],
        "mes_sources": {"trade_investment_agencies", "country_trade_organizations"},
    },
    "fundraising_investment": {
        "title": [
            r"\bventure\s+(capital|partner|associate|investor)\b",
            r"\bvc\s+(partner|associate|investor)\b",
            r"\bangel\s+investor\b",
            r"\binvestment\s+(associate|partner|director|principal|officer|manager)\b",
            r"\b(general|limited|managing|venture)\s+partner\b",
            r"\bhead\s+of\s+(venture|investment|capital)\b",
            r"\bportfolio\s+(manager|director)\b",
            r"\bfund\s+(manager|director|partner)\b",
            r"\b(seed|series\s+a)\s+investor\b",
            r"\bgrowth\s+investor\b",
        ],
        "company": [
            r"\bventures\b", r"\bcapital\b", r"\bblackbird\b", r"\bantler\b",
            r"\bairtree\b", r"\bsquare\s*peg\b", r"\bking\s*river\b",
            r"\baussie\s+angels\b", r"\binnovation\s+bay\b",
            r"\bh2\s+ventures\b", r"\btelstra\s+ventures\b",
            r"\bouture\s+capital\b", r"\bsixty\s+two\b",
            r"\bsavi\s+venture\b", r"\bm12\b",
            # Generic VC/PE/family office name fallbacks
            r"\bprivate\s+equity\b", r"\bfamily\s+office\b",
            r"\bvc\s+partners\b", r"\binvest(ment)?\s+(group|partners|fund)\b",
        ],
        "industry": ["venture capital & private equity",
                     "venture capital and private equity",
                     "investment management", "investment banking",
                     "private equity"],
    },
    "accounting_tax": {
        "title": [
            r"\baccountant\b",
            r"\b(senior\s+|managing\s+)?tax\s+(partner|advisor|adviser|director|manager|consultant)\b",
            r"\bauditor?\b", r"\baudit\s+(partner|director|manager)\b",
            r"\b(financial|finance)\s+controller\b",
            r"\b(senior|director).{0,15}\b(tax|audit|accounting)\b",
            r"\bcfo\b.{0,20}\b(advisory|consulting)\b",
            r"\btransfer\s+pricing\b",
        ],
        "company": [
            r"\bdeloitte\b", r"\bpwc\b", r"\bkpmg\b", r"\b(ey|ernst\s*&?\s*young)\b",
            r"\bgrant\s+thornton\b", r"\bbdo\b", r"\bpitcher\s+partners\b",
            r"\bfullstack\s+advisory\b", r"\bstandard\s+ledger\b",
            r"\bsleek\b",  # Sleek does accounting too
            r"\bmazars\b", r"\bnexia\b", r"\brsm\b", r"\bmoore\s+stephens\b",
        ],
        "industry": ["accounting", "tax services", "professional services - accounting"],
    },
    "legal": {
        "title": [
            r"\blawyer\b", r"\bsolicitor\b", r"\battorney\b", r"\bbarrister\b",
            r"\b(general|legal|chief)\s+counsel\b",
            r"\b(senior|managing)\s+(legal\s+)?associate\b",
            r"\b(legal|law)\s+partner\b",
            r"\bpartner\b.{0,30}\b(law|legal)\b",
            r"\bintellectual\s+property\s+(lawyer|partner)\b",
        ],
        "company": [
            r"\ballens\b", r"\bgilbert\s*\+\s*tobin\b", r"\bgtlaw\b",
            r"\bherbert\s+smith\b", r"\bking\s*&?\s*wood\b",
            r"\bminterellison\b", r"\bdla\s+piper\b",
            r"\bhall\s*(&|and)?\s*wilcox\b", r"\blander\s*(&|and)?\s*rogers\b",
            r"\bgadens\b", r"\blegalvision\b", r"\blawpath\b",
            r"\baddisons\b", r"\bdentons\b", r"\bk\s*&\s*l\s+gates\b",
            r"\bpinsent\s+masons\b", r"\bclayton\s+utz\b",
            r"\bbaker\s+mckenzie\b", r"\bnorton\s+rose\b",
            r"\bashurst\b", r"\blinklaters\b", r"\bcounsel\s+house\b",
        ],
        "industry": ["law practice", "legal services"],
    },
    "hr_recruitment_talent": {
        "title": [
            r"\b(head|director|chief|vp|founder).{0,25}\b(people|talent|hr|human\s+resources)\b",
            r"\brecruiter\b", r"\b(senior|principal)?\s*talent\s+acquisition\b",
            r"\bexecutive\s+search\b", r"\bstaffing\b",
            r"\b(people|talent)\s+ops\b", r"\bhead\s+of\s+culture\b",
        ],
        "company": [
            r"\bhays\b", r"\bhalcyon\s+knights\b", r"\bpolyglot\b",
            r"\bcloud\s+recruit\b", r"\btransearch\b",
            r"\brobert\s+walters\b", r"\bmichael\s+page\b",
            r"\bpage\s+executive\b", r"\bcornerstone\s+search\b",
            r"\bwhybrows\b", r"\blotus\s+people\b", r"\btalentbridge\b",
            r"\bspencer\s+stuart\b", r"\bkorn\s+ferry\b", r"\brandstad\b",
            # Generic recruitment-firm-name fallback
            r"\brecruit(ment)?\b", r"\bsearch\s+(group|partners|associates)\b",
            r"\btalent\b.{0,15}(group|partners|consulting)\b",
        ],
        "industry": ["human resources", "staffing and recruiting",
                     "professional training and coaching"],
    },
    "marketing_pr_comms": {
        "title": [
            r"\b(head|director|chief|vp|founder).{0,25}\b(marketing|communications|brand|comms|pr)\b",
            r"\b(content|digital|growth|product)\s+marketing\s+(director|manager|head|lead)\b",
            r"\bpublic\s+(relations|affairs)\b",
            r"\bmedia\s+relations\b", r"\bcorporate\s+affairs\b",
            r"\bbrand\s+(strategist|director|manager)\b",
            r"\bcmo\b", r"\bchief\s+marketing\s+officer\b",
        ],
        "company": [
            r"\bthrive\s+pr\b", r"\bbench\s+pr\b", r"\bgracosway\b",
            r"\bhawker\s+britton\b", r"\btg\s+public\s+affairs\b",
            r"\bsodali\b", r"\bredhill\b", r"\bellis\s+jones\b",
            r"\bgranton\b", r"\bliquid\s+digital\b", r"\bjaywing\b",
            r"\bedelman\b", r"\bweber\s+shandwick\b", r"\bogilvy\b",
            r"\bbcw\b", r"\bfgs\s+global\b", r"\bbrunswick\s+group\b",
        ],
        "industry": ["marketing and advertising", "marketing & advertising",
                     "public relations and communications", "advertising services",
                     "broadcast media", "communications"],
    },
    "technology_product": {
        "title": [
            r"\bcto\b", r"\bchief\s+technology\s+officer\b",
            r"\b(head|director|vp|svp).{0,25}\b(engineering|product|technology|tech)\b",
            r"\b(staff|principal|distinguished)\s+engineer\b",
            r"\bchief\s+product\s+officer\b", r"\bcpo\b",
            r"\b(senior\s+)?product\s+(manager|owner|lead|director)\b",
            r"\btech\s+lead\b", r"\bengineering\s+(director|manager|leader)\b",
            r"\bsoftware\s+engineering\s+(manager|director|leader)\b",
            r"\bsre\s+(lead|director)\b", r"\bdevops\b",
        ],
        "industry": ["computer software", "information technology and services",
                     "internet", "computer networking", "saas", "it services and it consulting"],
    },
    "ai_data": {
        "title": [
            r"\b(head|chief|director|vp).{0,25}\b(ai|artificial\s+intelligence|data\s+science|analytics|machine\s+learning|ml)\b",
            r"\bdata\s+scientist\b",
            r"\bmachine\s+learning\s+(engineer|scientist|lead|manager)\b",
            r"\b(senior\s+)?ml\s+engineer\b",
            r"\bai\s+(researcher|scientist|specialist)\b",
        ],
        "industry": ["artificial intelligence", "data infrastructure"],
    },
    "company_setup_eor": {
        "title": [
            r"\b(head|director|gm|general\s+manager).{0,25}\b(operations|expansion|payroll|hr\s+ops|workforce|people\s+operations)\b",
            r"\bemployer\s+of\s+record\b",
            r"\beor\s+specialist\b",
            r"\bglobal\s+expansion\s+(specialist|consultant|manager)\b",
        ],
        "company": [
            r"\bdeel\b", r"\bremote\.com\b", r"\bvelocity\s+global\b",
            r"\bpapaya\s+global\b", r"\brippling\b", r"\bsleek\b",
            r"\bacclime\b", r"\bvialto\b", r"\bemployment\s+hero\b",
            r"\bstandard\s+ledger\b", r"\bfullstack\s+advisory\b",
            r"\bairtasker\b",  # only if relevant
            r"\boyster\b", r"\bsafeguard\s+global\b", r"\bmultiplier\b",
        ],
    },
    "immigration_visa_mobility": {
        "title": [
            r"\bimmigration\s+(lawyer|advisor|adviser|consultant|specialist)\b",
            r"\bmigration\s+agent\b",
            r"\bglobal\s+mobility\s+(specialist|consultant|director|manager|partner)\b",
            r"\bvisa\s+(consultant|specialist)\b",
            r"\brelocation\s+(consultant|specialist|director|manager)\b",
            r"\bexpat(riate)?\s+(specialist|services)\b",
        ],
        "company": [
            r"\babsolute\s+immigration\b", r"\bpendlebury\b",
            r"\btechvisa\b", r"\bvisa\s+executive\b",
            r"\bewr\b", r"\belite\s+woodhams\b",
            r"\bfragomen\b", r"\bnewland\s+chase\b",
        ],
    },
    "compliance_risk": {
        "title": [
            r"\b(head|chief|director|vp).{0,20}\b(compliance|risk|regulatory|risk\s+&?\s+compliance)\b",
            r"\bquality\s+(assurance|control)\s+(manager|director|head)\b",
            r"\bregulatory\s+(affairs|specialist|director|manager)\b",
            r"\bchief\s+risk\s+officer\b", r"\bcro\b.{0,15}risk",
            r"\bcompliance\s+officer\b",
            r"\bdata\s+protection\s+officer\b", r"\bdpo\b",
            r"\bprivacy\s+(officer|counsel|specialist)\b",
        ],
        "industry": ["regulatory affairs", "risk management"],
    },
    "sales_gtm_leadgen": {
        "title": [
            r"\b(head|director|vp|svp|chief).{0,15}\b(sales|revenue|growth|business\s+development|partnerships|alliances|commercial)\b",
            r"\bcro\b.{0,20}(sales|revenue|commercial)?\b",
            r"\bchief\s+(revenue|commercial|sales|growth)\s+officer\b",
            r"\bsales\s+(director|leader|leadership|head)\b",
            r"\b(channel|alliance|strategic)\s+partnerships?\b",
            r"\bcommercial\s+(director|head|leader)\b",
            r"\bfractional\s+(cro|cmo|sales)\b",
            r"\b(senior\s+)?account\s+executive\b",
            r"\bnew\s+business\s+(director|manager|head)\b",
            r"\bbdr\s+(manager|director|lead)\b",
            r"\bsdr\s+(manager|director|lead)\b",
            r"\bsales\s+enablement\b",
            r"\brevops\b", r"\brevenue\s+operations\b",
        ],
    },
    "market_entry_specialist": {
        "title": [
            r"\bmarket\s+entry\s+(manager|specialist|director|lead)\b",
            r"\b(international|global|apac|anz)\s+expansion\b",
            r"\bgo[\s-]to[\s-]market\s+(strategist|director|lead)\b",
            r"\bcross-?border\s+(specialist|consultant)\b",
            r"\bregional\s+(launch|expansion)\s+(director|lead|manager)\b",
        ],
    },
    "market_research_strategy": {
        "title": [
            r"\bstrategy\s+consultant\b",
            r"\b(head|director|chief).{0,15}\bstrategy\b",
            r"\bmarket\s+research\b",
            r"\bstrategic\s+planning\s+(director|head|manager)\b",
            r"\bmanagement\s+consultant\b",
        ],
        "company": [
            r"\bmckinsey\b", r"\bbcg\b", r"\bbain\b",
            r"\baccenture\s+strategy\b", r"\bdeloitte\s+strategy\b",
            r"\boliver\s+wyman\b", r"\bey\s+parthenon\b",
            r"\bb2b\s+international\b", r"\bsis\s+international\b",
            r"\bstrategy\s+&\b",
        ],
        "industry": ["management consulting", "market research", "research"],
    },
    "events_community": {
        "title": [
            r"\b(head|director|manager).{0,15}\b(community|events|conference|programs?)\b",
            r"\bcommunity\s+(lead|manager|director)\b",
            r"\bevent\s+(director|producer|organiser|organizer)\b",
            r"\bprogram\s+(director|manager|lead)\b.{0,30}(accelerator|incubator|community)",
        ],
    },
    "logistics_freight_customs": {
        "title": [
            r"\b(logistics|freight|supply\s+chain|customs)\s+(manager|director|specialist|lead)\b",
            r"\b(head|chief).{0,15}\b(logistics|supply\s+chain|operations)\b.{0,15}(freight|customs|shipping)?",
            r"\bcustoms\s+broker\b",
        ],
        "company": [
            r"\bics\s+global\s+logistics\b", r"\btgl\b",
            r"\bthink\s+global\s+logistics\b", r"\bdb\s+schenker\b",
            r"\bdhl\b", r"\bfedex\b", r"\bups\b", r"\btoll\s+group\b",
            r"\blinfox\b",
        ],
        "industry": ["logistics and supply chain", "transportation/trucking/railroad",
                     "warehousing", "package/freight delivery"],
    },
    "translation_localisation": {
        "title": [
            r"\b(head|director|founder).{0,15}\b(translation|localisation|localization)\b",
            r"\btranslator\b",
            r"\binterpreter\b",
        ],
        "company": [
            r"\bamls\b", r"\bpolyglot\b", r"\bspeak\s+your\s+language\b",
            r"\baustralian\s+multilingual\b",
            r"\btranslation\b", r"\blocalisation\b", r"\blocalization\b",
        ],
    },
    "financial_services_fx_insurance": {
        "title": [
            r"\b(head|director|chief).{0,15}\b(fx|foreign\s+exchange|insurance|wealth|treasury|payments)\b",
            r"\binsurance\s+broker\b",
            r"\bwealth\s+manager\b", r"\bfinancial\s+(planner|advisor|adviser)\b",
            r"\btreasury\s+(director|manager)\b",
        ],
        "company": [
            r"\bofx\b", r"\bairwallex\b", r"\bconvera\b",
            r"\bmarsh\b", r"\bjudo\s+bank\b", r"\bnab\b",
            r"\bcommonwealth\s+bank\b", r"\bcommbank\b", r"\bwestpac\b",
            r"\banz\s+bank\b", r"\bmacquarie\s+(bank|group)\b",
            r"\bsuncorp\b", r"\biag\b", r"\baon\b", r"\bzurich\b",
            r"\bqbe\b",
        ],
        "industry": ["financial services", "banking", "insurance",
                     "investment banking", "capital markets"],
    },
    "startup_advisor": {
        "title": [
            r"\bmentor\s+in\s+residence\b",
            r"\bentrepreneur\s+in\s+residence\b",
            r"\b(head|program\s+(director|manager|lead)).{0,30}\b(accelerator|incubator|founders?)\b",
            r"\b(founder|co-?founder)\b.{0,20}(accelerator|incubator)",
        ],
        "company": [
            r"\bantler\b", r"\bstartmate\b", r"\btechstars\b",
            r"\by\s+combinator\b", r"\bfishburners\b",
            r"\bstone\s*&?\s*chalk\b", r"\bcicada\s+innovations\b",
            r"\bbluechilli\b", r"\bfounder\s+institute\b",
            r"\bmuru-?d\b", r"\bunited\s+founders\b", r"\bplus\s+eight\b",
            r"\briver\s+city\s+labs\b", r"\bhaymarket\s+hq\b",
            r"\bcremorne\s+digital\s+hub\b", r"\bmacquarie\s+university\s+incubator\b",
        ],
        "industry": ["venture capital & private equity"],  # secondary signal
    },
}

# Tier 3 sector keywords for sector_expert_* fallback
SECTOR_KEYWORDS = [
    "fintech", "healthtech", "medtech", "agritech", "cleantech", "climate tech",
    "edtech", "proptech", "insurtech", "regtech", "legaltech", "saas",
    "biotech", "life sciences", "pharmaceutical", "telecommunications",
    "telco", "energy", "renewables", "mining", "defence", "defense",
    "construction", "manufacturing", "automotive", "supply chain", "logistics",
    "banking", "fintech", "payments", "wealth management", "superannuation",
    "real estate", "hospitality", "retail", "consumer goods",
]

# Category priority — for tiebreakers (more specific wins)
CAT_PRIORITY = [
    "immigration_visa_mobility", "translation_localisation",
    "logistics_freight_customs",
    "company_setup_eor", "accounting_tax", "legal", "compliance_risk",
    "financial_services_fx_insurance",
    "fundraising_investment", "startup_advisor",
    "government_trade", "market_entry_specialist",
    "ai_data", "technology_product",
    "marketing_pr_comms", "hr_recruitment_talent",
    "events_community", "market_research_strategy",
    "sales_gtm_leadgen",
]

# Title is the strongest signal — what someone DOES dominates over where they
# work. Company match is a tiebreaker, not a primary signal. (Otherwise an
# engineer at NAB gets categorized as financial services, and a Head of
# Partnerships at Sleek gets categorized as accounting because Sleek's
# industry tag is 'Accounting'.)
WEIGHTS = {"title": 7, "company": 3, "industry": 2, "mes_source": 3}

def categorize(row) -> str:
    title    = (row.get("linkedinJobTitle") or "").lower()
    headline = (row.get("linkedinHeadline") or "").lower()
    desc     = (row.get("linkedinDescription") or "").lower()
    company  = (row.get("companyName") or "").lower()
    industry = (row.get("companyIndustry") or "").lower()
    mes_src  = row.get("mes_company_source") or ""

    title_text = title + " " + headline
    scores: dict[str, int] = {cat: 0 for cat in CAT_DEFS}

    for cat, defs in CAT_DEFS.items():
        for pat in defs.get("title", []):
            if re.search(pat, title_text):
                scores[cat] += WEIGHTS["title"]
        for pat in defs.get("company", []):
            if re.search(pat, company):
                scores[cat] += WEIGHTS["company"]
        for ind in defs.get("industry", []):
            if ind in industry:
                scores[cat] += WEIGHTS["industry"]
        if mes_src and mes_src in defs.get("mes_sources", set()):
            scores[cat] += WEIGHTS["mes_source"]

    best_score = max(scores.values())
    if best_score == 0:
        # Fallback: sector specialist if Tier 3 sector word in description/industry/company
        text_all = title_text + " " + desc + " " + industry + " " + company
        for sector in SECTOR_KEYWORDS:
            if sector in text_all:
                slug = sector.replace(" ", "_")
                return f"sector_expert_{slug}"
        return "general"

    # Pick highest, tiebreak by CAT_PRIORITY (lower index = wins)
    candidates = [c for c, s in scores.items() if s == best_score]
    if len(candidates) == 1:
        return candidates[0]
    candidates.sort(key=lambda c: CAT_PRIORITY.index(c) if c in CAT_PRIORITY else 999)
    return candidates[0]


print("Re-categorizing...")
df["mentor_category"] = df.apply(categorize, axis=1)

# ---------------------------------------------------------------------------
# Split financial_services_fx_insurance into fintech_founder vs traditional FS
# ---------------------------------------------------------------------------
TRADITIONAL_FS_COMPANIES = [
    "nab", "national australia bank", "commonwealth bank", "commbank",
    "westpac", "anz bank", "australia and new zealand banking",
    "macquarie bank", "macquarie group", "suncorp", "iag",
    "qbe", "aon", "marsh", "willis towers", "zurich australia",
    "allianz", "tal life", "amp limited", "asb bank",
    "bank of new zealand", "bnz", "kiwibank", "judo bank",
]
FINTECH_KEYWORDS = [
    "fintech", "insurtech", "regtech", "wealthtech", "neobank",
    "payments platform", "lending platform", "buy now pay later", "bnpl",
    "crypto", "web3", "stablecoin", "defi",
    "trading platform", "investment app", "robo[\\s-]advis",
    "digital bank", "challenger bank", "open banking",
    "embedded finance", "earned wage access",
]

def split_financial_services(row):
    if row["mentor_category"] != "financial_services_fx_insurance":
        return row["mentor_category"]
    title = (row.get("linkedinJobTitle") or "").lower()
    company = (row.get("companyName") or "").lower()
    headline = (row.get("linkedinHeadline") or "").lower()
    desc = (row.get("linkedinDescription") or "").lower()
    text = f"{title} {company} {headline} {desc}"
    # Explicit fintech keyword → fintech_founder regardless of role
    if any(re.search(rf"\b{kw}\b", text) for kw in FINTECH_KEYWORDS):
        return "fintech_founder"
    # Founder/co-founder at non-traditional company → fintech_founder
    if bool(row.get("is_founder")):
        if not any(t in company for t in TRADITIONAL_FS_COMPANIES):
            return "fintech_founder"
    return "financial_services"

df["mentor_category"] = df.apply(split_financial_services, axis=1)

# ---------------------------------------------------------------------------
# Consolidate sector_expert_* into single 'sector_expert' bucket; lift the
# vertical into a new sector_vertical column.
# ---------------------------------------------------------------------------
SECTOR_ALIASES = {"defense": "defence"}  # dedupe spelling variants
df["sector_vertical"] = df["mentor_category"].apply(
    lambda c: SECTOR_ALIASES.get(
        c.replace("sector_expert_", ""), c.replace("sector_expert_", "")
    ) if isinstance(c, str) and c.startswith("sector_expert_") else ""
)
df["mentor_category"] = df["mentor_category"].apply(
    lambda c: "sector_expert" if isinstance(c, str) and c.startswith("sector_expert_") else c
)

print("New mentor_category distribution (full pool):")
print(df["mentor_category"].value_counts().head(25))

# ---------------------------------------------------------------------------
# Apply Recipe Y filter: archetype + score >= 14, drop "general" bucket
# ---------------------------------------------------------------------------
qualified = df[
    df["any_archetype"]
    & (df["total_score"] >= 14)
    & (df["mentor_category"] != "general")
].copy()
print(f"\nRecipe Y qualified (after dropping 'general'): {len(qualified)}")
print("Recipe Y mentor_category distribution:")
print(qualified["mentor_category"].value_counts().head(25))

# Re-tier within the new pool (no C tier — floor is 14, not 8)
def assign_rating(score):
    if score >= 19:
        return "A"
    if score >= 17:
        return "B+"
    return "B"
qualified["rating"] = qualified["total_score"].apply(assign_rating)

# Build "archetypes" string for visibility
def archetype_label(row):
    a = []
    if row["archetype_advisor"]:        a.append("Advisor")
    if row["archetype_csuite_sector"]:  a.append("C-Suite/Founder")
    if row["archetype_trade_gov"]:      a.append("Trade/Gov")
    if row["archetype_intl_founder"]:   a.append("Intl Founder")
    return " + ".join(a)
qualified["archetypes"] = qualified.apply(archetype_label, axis=1)

# Sort
qualified = qualified.sort_values(
    by=["total_score", "rating"], ascending=[False, True]
).reset_index(drop=True)

# Final column order matching the spec
COLUMNS = [
    "fullName", "rating", "total_score", "archetypes", "reasoning",
    "seniority_score", "relevance_score", "industry_score", "network_score",
    "persona_fit_score", "intl_founder_score",
    "is_founder", "origin_country", "origin_signals",
    "persona_fit_label", "mentor_category", "sector_vertical",
    "linkedinJobTitle", "companyName", "companyIndustry", "location",
    "linkedinHeadline", "linkedinProfileUrl",
    "professionalEmail1", "personalEmail1",
    "linkedinDescription", "linkedinSkillsLabel",
    "keyword_matches",
    "already_in_google_sheet", "already_in_supabase",
    "is_mes_tagged", "mes_company_match", "mes_company_name",
    "mes_company_source", "mes_company_match_type",
    "linkedinFollowersCount", "previousCompanyName",
    "linkedinSchoolName", "linkedinPreviousSchoolName",
]

# Truncate description / skills to spec
qualified["linkedinDescription"] = qualified["linkedinDescription"].astype(str).str[:500]
qualified["linkedinSkillsLabel"] = qualified["linkedinSkillsLabel"].astype(str).str[:300]

main = qualified[COLUMNS].copy()

# ---------------------------------------------------------------------------
# Save CSV
# ---------------------------------------------------------------------------
csv_path = ROOT / "mes_mentor_candidates.csv"
main.to_csv(csv_path, index=False)
print(f"Saved CSV: {csv_path} ({len(main)} rows)")

# ---------------------------------------------------------------------------
# Build xlsx with multiple sheets
# ---------------------------------------------------------------------------
xlsx_path = ROOT / "mes_mentor_candidates.xlsx"

wb = Workbook()
wb.remove(wb.active)  # remove default sheet

# --- Sheet 1: Mentor Candidates ----------------------------------------------
ws_main = wb.create_sheet("Mentor Candidates")
for r in dataframe_to_rows(main, index=False, header=True):
    ws_main.append(r)

# Format header row
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="1F4E79")
for cell in ws_main[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws_main.row_dimensions[1].height = 32
ws_main.freeze_panes = "A2"

# Conditional fill by rating (column B)
fill_a   = PatternFill("solid", fgColor="C6EFCE")  # green
fill_bp  = PatternFill("solid", fgColor="FFEB9C")  # yellow
fill_b   = PatternFill("solid", fgColor="EAEAEA")  # light grey
rating_col_idx = COLUMNS.index("rating") + 1
for row_num in range(2, ws_main.max_row + 1):
    rating = ws_main.cell(row=row_num, column=rating_col_idx).value
    fill = {"A": fill_a, "B+": fill_bp, "B": fill_b}.get(rating)
    if fill:
        for col_num in range(1, len(COLUMNS) + 1):
            ws_main.cell(row=row_num, column=col_num).fill = fill

# Auto-width (with caps)
for col_idx, col_name in enumerate(COLUMNS, start=1):
    width_cap = 60 if col_name in {"reasoning", "linkedinHeadline",
                                   "linkedinDescription", "linkedinSkillsLabel",
                                   "keyword_matches", "linkedinJobTitle"} else 30
    max_len = max(len(col_name), main[col_name].astype(str).str.len().max() or 0)
    ws_main.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, width_cap)

# --- Sheet 2: Summary --------------------------------------------------------
ws_sum = wb.create_sheet("Summary")

def write_section(ws, title, rows, start_row):
    ws.cell(row=start_row, column=1, value=title).font = Font(bold=True, size=12)
    r = start_row + 1
    for label, value in rows:
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=value)
        r += 1
    return r + 1

next_row = 1
next_row = write_section(ws_sum, "Tier counts", [
    ("Total qualified (Recipe Y)", len(main)),
    ("Tier A (>=19)", (main["rating"] == "A").sum()),
    ("Tier B+ (17-18)", (main["rating"] == "B+").sum()),
    ("Tier B (14-16)", (main["rating"] == "B").sum()),
    ("Quick Wins (>=17)", (main["total_score"] >= 17).sum()),
], next_row)

next_row = write_section(ws_sum, "By archetype (overlapping)", [
    ("Active Advisor", qualified["archetype_advisor"].sum()),
    ("Experienced C-Suite/Founder", qualified["archetype_csuite_sector"].sum()),
    ("Trade & Government", qualified["archetype_trade_gov"].sum()),
    ("International Founder", qualified["archetype_intl_founder"].sum()),
], next_row)

# Mentor category
cat_counts = main["mentor_category"].value_counts()
next_row = write_section(ws_sum, "Mentor category (primary)",
    [(c, n) for c, n in cat_counts.items()], next_row)

# Persona fit label
pf_counts = main["persona_fit_label"].value_counts()
next_row = write_section(ws_sum, "Persona fit label",
    [(c, n) for c, n in pf_counts.items()], next_row)

# Intl founders by origin
intl = qualified[qualified["intl_founder_score"] >= 3]
intl_by_origin = intl.groupby("origin_country").size().sort_values(ascending=False)
next_row = write_section(ws_sum, "International founders by origin (>=3 score)",
    [(c, int(n)) for c, n in intl_by_origin.items()], next_row)

# Top industries
ind_counts = main["companyIndustry"].value_counts().head(15)
next_row = write_section(ws_sum, "Top 15 industries",
    [(c, int(n)) for c, n in ind_counts.items()], next_row)

# Top companies
co_counts = main["companyName"].value_counts().head(15)
next_row = write_section(ws_sum, "Top 15 companies",
    [(c, int(n)) for c, n in co_counts.items()], next_row)

# Locations
loc_counts = main["location"].value_counts().head(10)
next_row = write_section(ws_sum, "Top 10 locations",
    [(c, int(n)) for c, n in loc_counts.items()], next_row)

# MES tag / company match
next_row = write_section(ws_sum, "Ecosystem signal counts", [
    ("MES-tagged in network", main["is_mes_tagged"].sum()),
    ("MES company match (current/previous)", main["mes_company_match"].sum()),
], next_row)

ws_sum.column_dimensions["A"].width = 50
ws_sum.column_dimensions["B"].width = 20
ws_sum.freeze_panes = "A2"

# --- Sheet 3: International Founders ----------------------------------------
ws_intl = wb.create_sheet("International Founders")
intl_cols = ["fullName", "rating", "total_score", "archetypes",
             "intl_founder_score", "origin_country", "origin_signals",
             "linkedinJobTitle", "companyName", "location",
             "linkedinProfileUrl", "professionalEmail1", "personalEmail1",
             "reasoning"]
intl_df = qualified[qualified["intl_founder_score"] >= 3][intl_cols].copy()
intl_df = intl_df.sort_values(by=["origin_country", "total_score"],
                              ascending=[True, False]).reset_index(drop=True)

for r in dataframe_to_rows(intl_df, index=False, header=True):
    ws_intl.append(r)
for cell in ws_intl[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws_intl.row_dimensions[1].height = 32
ws_intl.freeze_panes = "A2"
for col_idx, col_name in enumerate(intl_cols, start=1):
    width_cap = 60 if col_name in {"reasoning", "linkedinJobTitle"} else 30
    max_len = max(len(col_name), intl_df[col_name].astype(str).str.len().max() or 0)
    ws_intl.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, width_cap)

# --- Sheet 4: Quick Wins ----------------------------------------------------
ws_qw = wb.create_sheet("Quick Wins")
qw_cols = ["fullName", "rating", "total_score", "archetypes",
           "mentor_category", "origin_country",
           "linkedinJobTitle", "companyName", "location",
           "linkedinProfileUrl", "professionalEmail1", "personalEmail1",
           "reasoning"]
qw_df = qualified[qualified["total_score"] >= 17][qw_cols].copy()
qw_df = qw_df.sort_values(by="total_score", ascending=False).head(50).reset_index(drop=True)

for r in dataframe_to_rows(qw_df, index=False, header=True):
    ws_qw.append(r)
for cell in ws_qw[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws_qw.row_dimensions[1].height = 32
ws_qw.freeze_panes = "A2"
for col_idx, col_name in enumerate(qw_cols, start=1):
    width_cap = 60 if col_name in {"reasoning", "linkedinJobTitle"} else 30
    max_len = max(len(col_name), qw_df[col_name].astype(str).str.len().max() or 0)
    ws_qw.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, width_cap)

# --- Sheet 5+: Per-archetype tabs -------------------------------------------
def add_archetype_sheet(name, mask, cols=None):
    cols = cols or ["fullName", "rating", "total_score", "archetypes",
                    "mentor_category", "linkedinJobTitle", "companyName",
                    "companyIndustry", "location", "linkedinProfileUrl",
                    "professionalEmail1", "personalEmail1", "reasoning"]
    sub = qualified[mask][cols].sort_values("total_score", ascending=False).reset_index(drop=True)
    ws = wb.create_sheet(name[:31])  # Excel sheet name limit is 31 chars
    for r in dataframe_to_rows(sub, index=False, header=True):
        ws.append(r)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"
    for col_idx, col_name in enumerate(cols, start=1):
        width_cap = 60 if col_name in {"reasoning", "linkedinJobTitle"} else 30
        max_len = max(len(col_name), sub[col_name].astype(str).str.len().max() or 0)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, width_cap)
    print(f"  Tab {name}: {len(sub)} rows")

print("Per-archetype tabs:")
add_archetype_sheet("By Archetype - Advisor", qualified["archetype_advisor"])
add_archetype_sheet("By Archetype - C-Suite Founder", qualified["archetype_csuite_sector"])
add_archetype_sheet("By Archetype - Trade Gov", qualified["archetype_trade_gov"])
add_archetype_sheet("By Archetype - Intl Founder", qualified["archetype_intl_founder"])

wb.save(xlsx_path)
print(f"\nSaved xlsx: {xlsx_path}")
print(f"  Total sheets: {len(wb.sheetnames)}")
print(f"  Sheets: {wb.sheetnames}")
print(f"\nFinal Recipe Y output: {len(main)} contacts")
print(f"  Tier A (>=19):  {(main['rating']=='A').sum()}")
print(f"  Tier B+ (17-18): {(main['rating']=='B+').sum()}")
print(f"  Tier B (14-16):  {(main['rating']=='B').sum()}")
print(f"  Quick Wins (>=17): {(main['total_score']>=17).sum()}")
print(f"  Intl Founders:   {len(intl_df)}")
