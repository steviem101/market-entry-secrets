"""Parse + profile the Startmate community-sourced ecosystem sheet (xlsx export).

Reads (not committed — PII-bearing, lives under the gitignored data/private/ path):
  data/private/startmate/startmate_community_sheet.xlsx
  (override with --xlsx; export the Google Sheet as .xlsx, all 7 tabs)

Writes:
  data/private/startmate/parsed_startmate_ecosystem.json   (candidate records, incl. emails)
  stdout: aggregate profiling report (no PII) — safe to paste into docs

No DB writes. Phase 1 (analysis / dry-run) of the Startmate ecosystem import
pipeline — see docs/data-analysis/startmate-ecosystem-sheet-analysis.md.

Usage:
  python3 scripts/parse_startmate_sheet.py [--xlsx PATH] [--self-test]

Requires: openpyxl (pip install openpyxl).
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from collections import Counter
from pathlib import Path
from urllib.parse import urlparse

REPO_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSX = REPO_ROOT / "data" / "private" / "startmate" / "startmate_community_sheet.xlsx"
OUTPUT_FILE = REPO_ROOT / "data" / "private" / "startmate" / "parsed_startmate_ecosystem.json"

SOURCE_NAME = "startmate_community_sheet"
SOURCE_URL = (
    "https://docs.google.com/spreadsheets/d/1lIK2Lji83-tO5DlgMHu7KI7WeVzPqL9fFJ2hrNuVaCg"
)

AU_CITIES = {
    "sydney", "melbourne", "brisbane", "perth", "adelaide", "canberra", "hobart",
    "darwin", "gold coast", "newcastle", "wollongong", "geelong", "byron bay",
    "parramatta", "liverpool", "shoalhaven", "wollondilly", "neerabup", "newman",
    "naarm", "queensland", "victoria", "nsw", "western australia", "tasmania",
    "south australia", "act", "national",
}
NZ_CITIES = {
    "auckland", "wellington", "christchurch", "hamilton", "dunedin", "tauranga",
    "queenstown",
}
US_CITIES = {
    "san francisco", "silicon valley", "new york", "nyc", "los angeles", "boston",
    "seattle", "austin", "chicago",
}

EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


# ---------------------------------------------------------------------------
# Normalization helpers (unit-tested via --self-test)
# ---------------------------------------------------------------------------

def clean(value) -> str | None:
    """Collapse whitespace; return None for empty cells."""
    if value is None:
        return None
    text = re.sub(r"\s+", " ", str(value)).strip()
    return text or None


def normalize_url(url: str | None) -> str | None:
    """Lowercase scheme/host, add https:// when missing, strip trailing slash."""
    url = clean(url)
    if not url:
        return None
    if not re.match(r"^https?://", url, re.I):
        url = "https://" + url
    parsed = urlparse(url)
    if not parsed.netloc or "." not in parsed.netloc:
        return None
    host = parsed.netloc.lower()
    path = parsed.path.rstrip("/")
    rebuilt = f"https://{host}{path}"
    if parsed.query:
        rebuilt += f"?{parsed.query}"
    return rebuilt


def extract_domain(url: str | None) -> str | None:
    """Registrable-ish domain: hostname lowercased, www. stripped."""
    normalized = normalize_url(url)
    if not normalized:
        return None
    host = urlparse(normalized).netloc
    return re.sub(r"^www\.", "", host)


def normalize_name(name: str | None) -> str | None:
    """Aggressive org-name key: lowercase, strip generic suffixes + punctuation."""
    name = clean(name)
    if not name:
        return None
    lowered = name.lower()
    lowered = re.sub(r"\((.*?)\)", " ", lowered)  # drop parentheticals
    lowered = re.sub(
        r"\b(ventures?|venture capital|capital|vc|partners?|fund|group|the|pty|ltd|inc|accelerator|program|programme)\b",
        " ",
        lowered,
    )
    lowered = re.sub(r"[^a-z0-9]+", "", lowered)
    return lowered or re.sub(r"[^a-z0-9]+", "", name.lower())


def classify_geography(country: str | None, city: str | None) -> str:
    """Bucket a row: australia | new_zealand | anz | us | global | unknown."""
    c = (clean(country) or "").lower()
    ci = (clean(city) or "").lower()
    if "australia" in c and ("new zealand" in c or "nz" in c.split("/") or "anz" in c):
        return "anz"
    if c in {"anz", "aus/nz", "australia / new zealand"}:
        return "anz"
    if "australia" in c or c in {"aus", "au"}:
        return "australia"
    if "new zealand" in c or c == "nz":
        return "new_zealand"
    if c in {"us", "usa", "united states"}:
        return "us"
    if c == "global":
        return "global"
    if ci:
        if ci in AU_CITIES or any(a in ci for a in AU_CITIES):
            return "australia"
        if ci in NZ_CITIES or any(n in ci for n in NZ_CITIES):
            return "new_zealand"
        if ci in US_CITIES or any(u in ci for u in US_CITIES):
            return "us"
    if c:
        # e.g. "Australia (QLD)", "Lower Hunter, Central Coast / Australia"
        if "australia" in c:
            return "australia"
        if "zealand" in c:
            return "new_zealand"
        return "other"
    return "unknown"


def classify_link(url: str | None) -> str:
    """Quality bucket for a URL: linkedin_profile | linkedin_company |
    linkedin_search | social | website | missing."""
    if not url:
        return "missing"
    lowered = url.lower()
    if "linkedin.com/search" in lowered or "google.com/search" in lowered:
        return "linkedin_search"
    if "linkedin.com/in/" in lowered:
        return "linkedin_profile"
    if "linkedin.com/company" in lowered:
        return "linkedin_company"
    if any(s in lowered for s in ("facebook.com", "instagram.com", "twitter.com", "x.com/")):
        return "social"
    return "website"


def dedupe_key(entity_type: str, name: str | None, website: str | None,
               geography: str, extra: str | None = None) -> str:
    """Deterministic candidate key: domain wins; else normalized name + geo (+extra)."""
    domain = extract_domain(website)
    if domain and "linkedin.com" not in domain and "facebook.com" not in domain:
        basis = f"{entity_type}:{domain}"
    else:
        basis = f"{entity_type}:{normalize_name(name) or ''}:{geography}:{normalize_name(extra) or ''}"
    return hashlib.md5(basis.encode()).hexdigest()[:16]


# ---------------------------------------------------------------------------
# Row -> candidate
# ---------------------------------------------------------------------------

def make_candidate(tab: str, row_index: int, entity_type: str, name: str | None,
                   raw: dict, *, website: str | None = None, country: str | None = None,
                   city: str | None = None, extra_key: str | None = None,
                   proposed_destination: str | None = None) -> dict | None:
    if not name:
        return None
    geography = classify_geography(country, city)
    website_norm = normalize_url(website)
    link_class = classify_link(website_norm)
    flags: list[str] = []
    if link_class == "missing":
        flags.append("missing_url")
    if link_class == "linkedin_search":
        flags.append("placeholder_search_link")
    if link_class == "social":
        flags.append("social_link_only")
    if geography == "unknown":
        flags.append("unknown_geography")
    email = clean(raw.get("email"))
    if email and not EMAIL_RE.match(email):
        flags.append("invalid_email")

    if geography in {"australia", "new_zealand", "anz"}:
        confidence = "high"
    elif geography in {"us", "global", "other"}:
        confidence = "medium"
    else:
        confidence = "low"
    if "placeholder_search_link" in flags or (
        "missing_url" in flags and not email and not clean(raw.get("contact"))
    ):
        confidence = "low"

    return {
        "source_name": SOURCE_NAME,
        "source_url": SOURCE_URL,
        "source_tab": tab,
        "row_index": row_index,
        "entity_type": entity_type,
        "name": clean(name),
        "geography": geography,
        "website": website_norm,
        "domain": extract_domain(website_norm),
        "link_class": link_class,
        "raw": {k: clean(v) if not isinstance(v, dict) else v for k, v in raw.items()},
        "proposed_destination": proposed_destination,
        "dedupe_key": dedupe_key(entity_type, name, website_norm, geography, extra_key),
        "validation_flags": flags,
        "confidence": confidence,
    }


# ---------------------------------------------------------------------------
# Tab parsers
# ---------------------------------------------------------------------------

def cellv(cell):
    return clean(cell.value)


def cell_url(cell):
    """Prefer the hyperlink target over display text (many cells just say 'Link')."""
    if cell.hyperlink is not None and cell.hyperlink.target:
        return cell.hyperlink.target.strip()
    return cellv(cell)


def parse_workbook(path: Path) -> list[dict]:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    candidates: list[dict] = []

    ws = wb["VCs"]
    for i, r in enumerate(ws.iter_rows(min_row=2), start=2):
        name = cellv(r[0])
        if not name:
            continue
        raw = {
            "full_name": name, "fund": cellv(r[1]), "linkedin": cell_url(r[2]),
            "job_title": cellv(r[4]), "city": cellv(r[5]), "country": cellv(r[6]),
            "sector_focus": cellv(r[7]),
        }
        cand = make_candidate(
            "VCs", i, "investor_person", name, raw,
            website=raw["linkedin"], country=raw["country"], city=raw["city"],
            extra_key=raw["fund"],
            proposed_destination="investors (contact enrichment or new angel row)",
        )
        if cand:
            candidates.append(cand)

    ws = wb["Accelerators"]
    for i, r in enumerate(ws.iter_rows(min_row=2), start=2):
        name = cellv(r[1])
        if not name:
            continue
        raw = {
            "status": cellv(r[0]), "name": name, "contact": cellv(r[2]),
            "city_country": cellv(r[3]), "investment": cellv(r[4]),
            "cohort_timing": cellv(r[5]), "focus": cellv(r[6]),
            "association": cellv(r[7]), "website": cell_url(r[8]),
        }
        cand = make_candidate(
            "Accelerators", i, "accelerator", name, raw,
            website=raw["website"], country=raw["city_country"],
            city=raw["city_country"],
            proposed_destination="innovation_ecosystem",
        )
        if cand:
            if (raw["status"] or "").lower() == "inactive":
                cand["validation_flags"].append("inactive")
                cand["confidence"] = "low"
            candidates.append(cand)

    ws = wb["Startup Newsletters "]
    for i, r in enumerate(ws.iter_rows(min_row=2), start=2):
        name = cellv(r[0])
        if not name:
            continue
        raw = {
            "name": name, "country": cellv(r[1]), "contact": cell_url(r[2]),
            "focus": cellv(r[3]), "link": cell_url(r[4]),
        }
        cand = make_candidate(
            "Startup Newsletters", i, "newsletter", name, raw,
            website=raw["link"], country=raw["country"],
            proposed_destination="content_items (curated guide) / KB source list",
        )
        if cand:
            candidates.append(cand)

    ws = wb["Coworking Spaces"]
    for i, r in enumerate(ws.iter_rows(min_row=2), start=2):
        name = cellv(r[0])
        if not name:
            continue
        raw = {
            "name": name, "country": cellv(r[1]), "city": cellv(r[2]),
            "suburb": cellv(r[3]), "contact": cell_url(r[4]), "cost": cellv(r[5]),
            "focus": cellv(r[6]), "website": cell_url(r[7]),
        }
        cand = make_candidate(
            "Coworking Spaces", i, "coworking_space", name, raw,
            website=raw["website"], country=raw["country"], city=raw["city"],
            proposed_destination="innovation_ecosystem (services: Co-working)",
        )
        if cand:
            candidates.append(cand)

    ws = wb["Student Societies"]
    for i, r in enumerate(ws.iter_rows(min_row=2), start=2):
        name = cellv(r[1])
        if not name:
            continue
        raw = {
            "active": cellv(r[0]), "name": name, "university": cellv(r[2]),
            "contact": cellv(r[3]), "role": cellv(r[4]), "email": cellv(r[5]),
            "website": cell_url(r[6]),
        }
        cand = make_candidate(
            "Student Societies", i, "student_society", name, raw,
            website=raw["website"], country=None, city=None,
            extra_key=raw["university"],
            proposed_destination="exclude (phase 1) / optional KB summary",
        )
        if cand:
            # Geography via university, not city columns.
            uni = (raw["university"] or "").lower()
            if any(k in uni for k in ("auckland", "otago", "canterbury", "wellington", "waikato")):
                cand["geography"] = "new_zealand"
            elif uni and "elon" not in uni:
                cand["geography"] = "australia"
            cand["validation_flags"] = [f for f in cand["validation_flags"] if f != "unknown_geography"]
            if not any(k in cand["name"].lower() for k in (
                "entrepreneur", "startup", "start-up", "innovat", "founder",
                "venture", "consult", "business", "fintech", "tech", "incubat",
            )):
                cand["validation_flags"].append("non_startup_society")
                cand["confidence"] = "low"
            candidates.append(cand)

    ws = wb["Workshop Hosts"]
    for i, r in enumerate(ws.iter_rows(min_row=3), start=3):  # 2 header rows
        name = cellv(r[2])
        if not name:
            continue
        raw = {
            "submitted_by": cellv(r[0]), "topic": cellv(r[1]), "name": name,
            "linkedin": cell_url(r[3]), "email": cellv(r[4]),
            "attended": cellv(r[5]), "comments": cellv(r[6]),
        }
        cand = make_candidate(
            "Workshop Hosts", i, "workshop_host", name, raw,
            website=raw["linkedin"], country="Australia",
            proposed_destination="exclude (phase 1) / directory_submissions if pursued",
        )
        if cand:
            candidates.append(cand)

    ws = wb["Podcasts"]
    for i, r in enumerate(ws.iter_rows(min_row=2), start=2):
        name = cellv(r[0])
        if not name:
            continue
        raw = {
            "name": name, "country": cellv(r[1]), "contact": cellv(r[2]),
            "focus": cellv(r[3]), "link": cell_url(r[4]),
        }
        cand = make_candidate(
            "Podcasts", i, "podcast", name, raw,
            website=raw["link"], country=raw["country"],
            proposed_destination="content_items (curated guide) / KB source list",
        )
        if cand:
            candidates.append(cand)

    return candidates


# ---------------------------------------------------------------------------
# Profiling report (aggregates only — no PII)
# ---------------------------------------------------------------------------

def profile(candidates: list[dict]) -> str:
    lines = ["# Startmate sheet profiling report", ""]
    by_tab: dict[str, list[dict]] = {}
    for c in candidates:
        by_tab.setdefault(c["source_tab"], []).append(c)

    lines.append(f"Total candidate rows: {len(candidates)}")
    dupes = Counter(c["dedupe_key"] for c in candidates)
    dupe_keys = {k for k, n in dupes.items() if n > 1}
    lines.append(f"Intra-sheet duplicate keys: {len(dupe_keys)}")
    lines.append("")

    for tab, rows in by_tab.items():
        lines.append(f"## {tab} — {len(rows)} rows")
        lines.append(f"- geography: {dict(Counter(r['geography'] for r in rows))}")
        lines.append(f"- confidence: {dict(Counter(r['confidence'] for r in rows))}")
        lines.append(f"- link quality: {dict(Counter(r['link_class'] for r in rows))}")
        flags = Counter(f for r in rows for f in r["validation_flags"])
        lines.append(f"- validation flags: {dict(flags)}")
        tab_dupes = [r["name"] for r in rows if r["dedupe_key"] in dupe_keys]
        if tab_dupes:
            lines.append(f"- duplicate-key names: {sorted(set(tab_dupes))}")
        lines.append("")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Self-test for the normalization helpers
# ---------------------------------------------------------------------------

def self_test() -> int:
    cases = [
        (normalize_url("gridakl.com"), "https://gridakl.com"),
        (normalize_url(" https://www.Startmate.com/ "), "https://www.startmate.com"),
        (normalize_url(None), None),
        (normalize_url("N/A"), None),
        (extract_domain("https://www.stoneandchalk.com.au/"), "stoneandchalk.com.au"),
        (extract_domain("not a url"), None),
        (normalize_name("Blackbird Ventures"), normalize_name("Blackbird")),
        (normalize_name("SquarePeg Capital"), normalize_name("Square Peg Capital")),
        (normalize_name("Plus Eight (Spacecubed)"), normalize_name("Plus Eight")),
        (classify_geography("Australia (QLD)", None), "australia"),
        (classify_geography("AUS/NZ", None), "anz"),
        (classify_geography(None, "Auckland"), "new_zealand"),
        (classify_geography("US", "San Francisco"), "us"),
        (classify_geography(None, None), "unknown"),
        (classify_link("https://www.linkedin.com/search/results/all/?keywords=x"), "linkedin_search"),
        (classify_link("https://www.linkedin.com/in/someone"), "linkedin_profile"),
        (classify_link("https://www.facebook.com/somesociety"), "social"),
        (classify_link(None), "missing"),
        # domain-based key beats name variants
        (
            dedupe_key("accelerator", "Startmate", "https://www.startmate.com/", "anz"),
            dedupe_key("accelerator", "Startmate Accelerator", "startmate.com", "australia"),
        ),
        # name+geo key when no domain
        (
            dedupe_key("coworking_space", "WorkLife", None, "australia"),
            dedupe_key("coworking_space", "worklife", None, "australia"),
        ),
    ]
    failures = [(i, got, want) for i, (got, want) in enumerate(cases) if got != want]
    for i, got, want in failures:
        print(f"FAIL case {i}: got {got!r}, want {want!r}", file=sys.stderr)
    print(f"self-test: {len(cases) - len(failures)}/{len(cases)} passed")
    return 1 if failures else 0


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--out", type=Path, default=OUTPUT_FILE)
    parser.add_argument("--self-test", action="store_true")
    args = parser.parse_args()

    if args.self_test:
        return self_test()

    if not args.xlsx.exists():
        print(
            f"Sheet not found: {args.xlsx}\n"
            "Export the Google Sheet as .xlsx (File > Download > Microsoft Excel) "
            "and place it at that path, or pass --xlsx.",
            file=sys.stderr,
        )
        return 1

    candidates = parse_workbook(args.xlsx)
    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(json.dumps(candidates, indent=1, ensure_ascii=False))
    print(profile(candidates))
    print(f"Wrote {len(candidates)} candidates to {args.out} (gitignored; contains PII).")
    return 0


if __name__ == "__main__":
    sys.exit(main())
